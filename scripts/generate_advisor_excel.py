import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ─── 顏色常數 ──────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")  # 深藍
HDR2_FILL = PatternFill("solid", fgColor="2E75B6")  # 中藍
HDR3_FILL = PatternFill("solid", fgColor="BDD7EE")  # 淺藍
ALT_FILL  = PatternFill("solid", fgColor="F2F7FB")  # 淡藍交替列
BEST_FILL = PatternFill("solid", fgColor="FFFF99")  # 黃色最佳
WARN_FILL = PatternFill("solid", fgColor="FFE0E0")  # 淡紅警示
GREEN_FILL= PatternFill("solid", fgColor="E2EFDA")  # 淡綠強調
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HDR2_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BOLD_FONT = Font(bold=True, name="Calibri", size=10)
NORM_FONT = Font(name="Calibri", size=10)
CTR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT_ALIGN = Alignment(horizontal="left",   vertical="center")
thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── 方法元資料解析 ──────────────────────────────────────────
METHOD_META = {
    "retrain_none":               ("Re-train",   "Old+New（合併）", "None"),
    "retrain_undersampling":      ("Re-train",   "Old+New（合併）", "Under-sampling"),
    "retrain_oversampling":       ("Re-train",   "Old+New（合併）", "Over-sampling(SMOTE)"),
    "retrain_hybrid":             ("Re-train",   "Old+New（合併）", "Hybrid(SMOTEENN)"),
    "finetune_none":              ("Fine-tune",  "Old→New（時序）", "None"),
    "finetune_undersampling":     ("Fine-tune",  "Old→New（時序）", "Under-sampling"),
    "finetune_oversampling":      ("Fine-tune",  "Old→New（時序）", "Over-sampling(SMOTE)"),
    "finetune_hybrid":            ("Fine-tune",  "Old→New（時序）", "Hybrid(SMOTEENN)"),
    "ensemble_old_3":             ("Ensemble",   "Old（舊資料池）",  "Under+Over+Hybrid"),
    "ensemble_new_3":             ("Ensemble",   "New（新資料池）",  "Under+Over+Hybrid"),
    "ensemble_all_6":             ("Ensemble",   "Old+New（雙池）",  "All 6 models"),
    "ensemble_2":                 ("Ensemble",   "Old+New（雙池）",  "Hybrid×2"),
    "ensemble_3_type_a":          ("Ensemble",   "Old+New（雙池）",  "Under+Over(old)+Hybrid(new)"),
    "ensemble_3_type_b":          ("Ensemble",   "Old+New（雙池）",  "Hybrid(old)+Over+Hybrid(new)"),
    "ensemble_4":                 ("Ensemble",   "Old+New（雙池）",  "Under+Over(×2)+Hybrid"),
    "ensemble_5":                 ("Ensemble",   "Old+New（雙池）",  "5 models（去New-Under）"),
    "DES_KNORAE":                 ("DES",        "Old+New（雙池）",  "All 6 models"),
    "DES_baseline":               ("DES",        "Old+New（雙池）",  "All 6 models"),
    "DES_time_weighted":          ("DES+TW",     "Old+New（雙池）",  "All 6 models（時間加權）"),
    "DES_minority_weighted":      ("DES+MW",     "Old+New（雙池）",  "All 6 models（少數類加權）"),
    "DES_combined":               ("DES+TW+MW",  "Old+New（雙池）",  "All 6 models（雙重加權）"),
}
def get_meta(m):
    if m in METHOD_META:
        return METHOD_META[m]
    return (m, "-", "-")

# ─── 工具函數 ──────────────────────────────────────────────
def read_csv_safe(path):
    try:
        df = pd.read_csv(path)
        if "Unnamed: 0" in df.columns:
            df = df.rename(columns={"Unnamed: 0": "method"})
        return df
    except:
        return pd.DataFrame()

def norm_cols(df):
    rename = {
        "Type1_Error(FPR)": "Type1_Error",
        "Type2_Error(FNR)": "Type2_Error",
        "G_Mean": "G_Mean",
    }
    return df.rename(columns=rename)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c

def write_header_row(ws, row, cols, font=HDR2_FONT, fill=HDR2_FILL, height=30):
    for i, col in enumerate(cols, 1):
        set_cell(ws, row, i, col, font=font, fill=fill, align=CTR_ALIGN, border=THIN_BORDER)
    ws.row_dimensions[row].height = height

def write_data_rows(ws, start_row, df, col_order, metric_cols, best_rows_idx=None):
    """Write data rows with alternating color and highlight best metric per column."""
    for r_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = start_row + r_idx
        fill = ALT_FILL if r_idx % 2 == 1 else None
        for c_idx, col in enumerate(col_order, 1):
            val = row.get(col, "")
            if pd.isna(val): val = "-"
            nf = None
            if col in metric_cols and isinstance(val, (int, float)) and val != "-":
                nf = "0.0000"
            set_cell(ws, excel_row, c_idx, val,
                     font=NORM_FONT, fill=fill, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
    return start_row + len(df)

def add_color_scale(ws, col_letter, start_row, end_row, reverse=False):
    """Add green-yellow-red color scale to a column range."""
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    if reverse:  # lower = better (e.g., Type1_Error)
        rule = ColorScaleRule(
            start_type="min",  start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="F8696B")
    else:        # higher = better (AUC, F1, Recall)
        rule = ColorScaleRule(
            start_type="min",  start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="63BE7B")
    ws.conditional_formatting.add(rng, rule)

def freeze_and_autofit(ws, freeze="A3"):
    ws.freeze_panes = freeze
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 32)

# ─── Sheet 標題列 ──────────────────────────────────────────
METRICS_EN  = ["AUC", "F1", "Precision", "Recall", "G_Mean", "Type1_Error", "Type2_Error"]
METRICS_ZH  = ["AUC\n(整體辨別力)", "F1\n(少數類)", "Precision\n(少數類精準率)",
               "Recall\n(少數類召回率)★", "G-Mean\n(平衡指標)",
               "Type1_Error\n(FPR 誤判率)★", "Type2_Error\n(FNR 漏報率)"]
BASE_COLS   = ["方法名稱", "策略", "訓練資料", "採樣策略"]
ALL_COLS_ZH = BASE_COLS + METRICS_ZH
ALL_COLS_EN = ["method", "strategy", "train_data", "sampling"] + METRICS_EN

def enrich_df(df):
    df = norm_cols(df.copy())
    strat, train, samp = [], [], []
    for m in df["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    df.insert(1, "strategy",   strat)
    df.insert(2, "train_data", train)
    df.insert(3, "sampling",   samp)
    for col in METRICS_EN:
        if col not in df.columns:
            df[col] = np.nan
    return df

def write_dataset_sheet(wb, sheet_name, df, title, note=""):
    ws = wb.create_sheet(sheet_name)
    # 大標題
    ws.merge_cells(f"A1:{get_column_letter(len(ALL_COLS_ZH))}1")
    set_cell(ws, 1, 1, title, font=Font(bold=True, color="FFFFFF", name="Calibri", size=13),
             fill=HDR_FILL, align=CTR_ALIGN)
    ws.row_dimensions[1].height = 36
    # 欄位標題
    write_header_row(ws, 2, ALL_COLS_ZH, height=45)
    # 資料
    df = enrich_df(df)
    metric_cols_set = set(METRICS_EN)
    data_start = 3
    for r_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = data_start + r_idx
        row_fill = ALT_FILL if r_idx % 2 == 1 else None
        for c_idx, col in enumerate(ALL_COLS_EN, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if col in metric_cols_set and isinstance(val, float) else None
            f  = NORM_FONT
            if col == "method": f = Font(bold=True, name="Calibri", size=10)
            set_cell(ws, excel_row, c_idx, val,
                     font=f, fill=row_fill, align=CTR_ALIGN if col != "方法名稱" else LFT_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws.row_dimensions[excel_row].height = 20
    # 色階 (欄位 5=AUC 6=F1 7=Prec 8=Recall 9=GMean 10=T1E 11=T2E)
    data_end = data_start + len(df) - 1
    for ci, col in enumerate(METRICS_EN, 5):
        cl = get_column_letter(ci)
        reverse = col in ("Type1_Error", "Type2_Error")
        add_color_scale(ws, cl, data_start, data_end, reverse=reverse)
    # 備注
    if note:
        note_row = data_end + 2
        ws.merge_cells(f"A{note_row}:{get_column_letter(len(ALL_COLS_ZH))}{note_row}")
        set_cell(ws, note_row, 1, f"※ {note}",
                 font=Font(italic=True, color="595959", name="Calibri", size=9), align=LFT_ALIGN)
    freeze_and_autofit(ws)
    return ws

# ═══════════════════════════════════════════════════════════
#  主程式
# ═══════════════════════════════════════════════════════════
BASE = r"c:\0_git workspace\Continual-Imbalance-Ensemble\results"
wb = Workbook()
wb.remove(wb.active)  # 移除預設 Sheet

# ─────────────────────────────────────────────────────────
# Sheet 0: 總覽 (Summary)
# ─────────────────────────────────────────────────────────
summary_df = read_csv_safe(os.path.join(BASE, "summary_all_datasets_detailed.csv"))
summary_df = norm_cols(summary_df)
ws0 = wb.create_sheet("00_總覽_Summary")
SUMM_COLS_ZH = ["資料集", "方法名稱", "策略", "訓練資料", "採樣策略"] + METRICS_ZH
SUMM_COLS_EN = ["dataset", "method", "strategy", "train_data", "sampling"] + METRICS_EN
ws0.merge_cells(f"A1:{get_column_letter(len(SUMM_COLS_ZH))}1")
set_cell(ws0, 1, 1, "研究結果總覽 — 全資料集 × 全方法（2026-03-02）",
         font=Font(bold=True, color="FFFFFF", name="Calibri", size=14),
         fill=HDR_FILL, align=CTR_ALIGN)
ws0.row_dimensions[1].height = 38
write_header_row(ws0, 2, SUMM_COLS_ZH, height=50)
# 按 dataset 排序
ds_order = {"bankruptcy": 0, "stock": 1, "medical": 2}
summary_df["_sort"] = summary_df["dataset"].map(ds_order).fillna(9)
summary_df = summary_df.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
strat, train, samp = [], [], []
for m in summary_df["method"]:
    s, t, sa = get_meta(m)
    strat.append(s); train.append(t); samp.append(sa)
summary_df.insert(2, "strategy",   strat)
summary_df.insert(3, "train_data", train)
summary_df.insert(4, "sampling",   samp)
for col in METRICS_EN:
    if col not in summary_df.columns: summary_df[col] = np.nan
DS_FILLS = {"bankruptcy": PatternFill("solid", fgColor="FFF2CC"),
            "stock":       PatternFill("solid", fgColor="E2EFDA"),
            "medical":     PatternFill("solid", fgColor="FCE4D6")}
data_start = 3
for r_idx, (_, row) in enumerate(summary_df.iterrows()):
    er = data_start + r_idx
    ds = str(row.get("dataset","")).lower()
    row_fill = DS_FILLS.get(ds, None)
    for c_idx, col in enumerate(SUMM_COLS_EN, 1):
        val = row.get(col, "")
        if isinstance(val, float) and np.isnan(val): val = "-"
        nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
        set_cell(ws0, er, c_idx, val, font=NORM_FONT, fill=row_fill,
                 align=CTR_ALIGN, border=THIN_BORDER, number_format=nf)
    ws0.row_dimensions[er].height = 18
data_end = data_start + len(summary_df) - 1
for ci, col in enumerate(METRICS_EN, 6):
    cl = get_column_letter(ci)
    reverse = col in ("Type1_Error", "Type2_Error")
    add_color_scale(ws0, cl, data_start, data_end, reverse=reverse)
note_r = data_end + 2
ws0.merge_cells(f"A{note_r}:{get_column_letter(len(SUMM_COLS_ZH))}{note_r}")
set_cell(ws0, note_r, 1,
         "★ Recall（少數類召回率）與 Type1_Error（誤判率）為本研究重點關注指標；色階：綠=較優，紅=較差",
         font=Font(italic=True, color="595959", size=9), align=LFT_ALIGN)
ws0.merge_cells(f"A{note_r+1}:{get_column_letter(len(SUMM_COLS_ZH))}{note_r+1}")
set_cell(ws0, note_r+1, 1, "Row color: 黃=Bankruptcy / 綠=Stock / 橘=Medical",
         font=Font(italic=True, color="595959", size=9), align=LFT_ALIGN)
freeze_and_autofit(ws0)

# ─────────────────────────────────────────────────────────
# Sheets 1–3: 各資料集 Baseline
# ─────────────────────────────────────────────────────────
# Bankruptcy Baseline
bkb = read_csv_safe(os.path.join(BASE, "baseline", "bankruptcy_baseline_results.csv"))
write_dataset_sheet(wb, "01_Bankruptcy_Baseline", bkb,
    "Bankruptcy — Baseline 結果（Re-train vs Fine-tune × 採樣策略）",
    note="↑ AUC/F1/Recall 越高越好；↓ Type1_Error（誤判率）越低越好。Fine-tune 保留時序訓練順序，理論上可減少遺忘。")

# Stock Baseline
stb = read_csv_safe(os.path.join(BASE, "stock", "stock_baseline_results.csv"))
write_dataset_sheet(wb, "02_Stock_Baseline", stb,
    "Stock (SPX) — Baseline 結果（Re-train vs Fine-tune × 採樣策略）",
    note="Stock 資料集整體 AUC≈0.55，接近隨機（弱式效率市場）。F1=0 代表閾值0.5 下模型不預測任何崩跌。")

# Medical Baseline
mdb = read_csv_safe(os.path.join(BASE, "medical", "medical_baseline_results.csv"))
write_dataset_sheet(wb, "03_Medical_Baseline", mdb,
    "Medical — Baseline 結果（Re-train vs Fine-tune × 採樣策略）",
    note="Re-train 的 Recall 雖高，但 Type1_Error≈0.37（37% 正常病人被誤判高危）；Fine-tune+Hybrid 的 Type1_Error 僅 2.75%。")

# ─────────────────────────────────────────────────────────
# Sheets 4–6: 各資料集 Ensemble
# ─────────────────────────────────────────────────────────
bke = read_csv_safe(os.path.join(BASE, "ensemble", "bankruptcy_ensemble_results.csv"))
write_dataset_sheet(wb, "04_Bankruptcy_Ensemble", bke,
    "Bankruptcy — Ensemble 結果（雙池集成 8 種組合）",
    note="Ensemble New-3 = 只用新資料池 3 個子模型（Under+Over+Hybrid）。理論上新資料池反映最新概念分布。")

ste = read_csv_safe(os.path.join(BASE, "ensemble", "stock_ensemble_results.csv"))
write_dataset_sheet(wb, "05_Stock_Ensemble", ste,
    "Stock (SPX) — Ensemble 結果（雙池集成）",
    note="多數方法 F1=0（predefault 閾值下不預測任何崩跌）；Ensemble Old-3 Recall=0.72 代表偏保守策略（寧多報）。")

mde = read_csv_safe(os.path.join(BASE, "ensemble", "medical_ensemble_results.csv"))
write_dataset_sheet(wb, "06_Medical_Ensemble", mde,
    "Medical — Ensemble 結果（雙池集成）",
    note="Ensemble New-3 在 AUC 與 Type1_Error 均最佳；但 Recall 偏低（0.126），需與臨床需求取捨。")

# ─────────────────────────────────────────────────────────
# Sheet 7: DES Advanced (3 資料集合併)
# ─────────────────────────────────────────────────────────
des_dfs = []
for ds, fpath in [
    ("Bankruptcy", os.path.join(BASE, "des_advanced", "bankruptcy_des_advanced_comparison.csv")),
    ("Stock",      os.path.join(BASE, "des_advanced", "stock_des_advanced_comparison.csv")),
    ("Medical",    os.path.join(BASE, "des_advanced", "medical_des_advanced_comparison.csv")),
]:
    df = read_csv_safe(fpath)
    if not df.empty:
        df.insert(0, "dataset", ds)
        des_dfs.append(df)
if des_dfs:
    des_all = pd.concat(des_dfs, ignore_index=True)
    DCOLS_ZH = ["資料集", "方法名稱", "策略", "訓練資料", "採樣策略"] + METRICS_ZH
    DCOLS_EN = ["dataset", "method", "strategy", "train_data", "sampling"] + METRICS_EN
    ws7 = wb.create_sheet("07_DES_Advanced")
    ws7.merge_cells(f"A1:{get_column_letter(len(DCOLS_ZH))}1")
    set_cell(ws7, 1, 1, "DES 進階比較 — KNORA-E / Time-weighted / Minority-weighted / Combined（3 資料集）",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws7.row_dimensions[1].height = 36
    write_header_row(ws7, 2, DCOLS_ZH, height=48)
    des_all = norm_cols(des_all)
    strat, train, samp = [], [], []
    for m in des_all["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    des_all.insert(2, "strategy",   strat)
    des_all.insert(3, "train_data", train)
    des_all.insert(4, "sampling",   samp)
    for col in METRICS_EN:
        if col not in des_all.columns: des_all[col] = np.nan
    ds_row_fills = {"Bankruptcy": PatternFill("solid", fgColor="FFF2CC"),
                    "Stock":       PatternFill("solid", fgColor="E2EFDA"),
                    "Medical":     PatternFill("solid", fgColor="FCE4D6")}
    ds_r = 3
    for r_idx, (_, row) in enumerate(des_all.iterrows()):
        er = ds_r + r_idx
        rf = ds_row_fills.get(str(row.get("dataset","")), None)
        for c_idx, col in enumerate(DCOLS_EN, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
            set_cell(ws7, er, c_idx, val, font=NORM_FONT, fill=rf, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws7.row_dimensions[er].height = 18
    de = ds_r + len(des_all) - 1
    for ci, col in enumerate(METRICS_EN, 6):
        add_color_scale(ws7, get_column_letter(ci), ds_r, de, col in ("Type1_Error","Type2_Error"))
    freeze_and_autofit(ws7)

# ─────────────────────────────────────────────────────────
# Sheet 8: 採樣策略比較 Sampling Comparison
# ─────────────────────────────────────────────────────────
# 抓 Baseline 全部方法，3 資料集合併，重點看 None/Under/Over/Hybrid
samp_dfs = []
for ds, bpath, epath in [
    ("Bankruptcy", os.path.join(BASE,"baseline","bankruptcy_baseline_results.csv"),
                   os.path.join(BASE,"ensemble","bankruptcy_ensemble_results.csv")),
    ("Stock",      os.path.join(BASE,"stock","stock_baseline_results.csv"),
                   os.path.join(BASE,"ensemble","stock_ensemble_results.csv")),
    ("Medical",    os.path.join(BASE,"medical","medical_baseline_results.csv"),
                   os.path.join(BASE,"ensemble","medical_ensemble_results.csv")),
]:
    b = read_csv_safe(bpath)
    e = read_csv_safe(epath)
    for df in [b, e]:
        if not df.empty:
            df.insert(0, "dataset", ds)
            samp_dfs.append(df)
if samp_dfs:
    samp_all = pd.concat(samp_dfs, ignore_index=True)
    samp_all = norm_cols(samp_all)
    strat, train, samp = [], [], []
    for m in samp_all["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    samp_all.insert(2, "strategy",   strat)
    samp_all.insert(3, "train_data", train)
    samp_all.insert(4, "sampling",   samp)
    for col in METRICS_EN:
        if col not in samp_all.columns: samp_all[col] = np.nan
    SCOLS_ZH = ["資料集", "方法名稱", "策略", "訓練資料", "採樣策略"] + METRICS_ZH
    SCOLS_EN = ["dataset", "method", "strategy", "train_data", "sampling"] + METRICS_EN
    ws8 = wb.create_sheet("08_採樣策略比較")
    ws8.merge_cells(f"A1:{get_column_letter(len(SCOLS_ZH))}1")
    set_cell(ws8, 1, 1, "採樣策略比較 — None vs Under vs Over vs Hybrid（Baseline + Ensemble 全部方法）",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws8.row_dimensions[1].height = 36
    # 分組子標題
    groups = [("None（無採樣）","None"), ("Under-sampling","Under-sampling"),
              ("Over-sampling(SMOTE)","Over-sampling(SMOTE)"), ("Hybrid(SMOTEENN)","Hybrid(SMOTEENN)")]
    current_row = 2
    ds_fills = {"Bankruptcy": PatternFill("solid","FFF2CC"), "Stock": PatternFill("solid","E2EFDA"),
                "Medical": PatternFill("solid","FCE4D6")}
    for grp_name, grp_key in groups:
        sub = samp_all[samp_all["sampling"].str.contains(grp_key, na=False, regex=False)] if grp_key != "None" else samp_all[samp_all["sampling"] == "None"]
        if sub.empty and grp_key == "None":
            sub = samp_all[samp_all["strategy"].isin(["Re-train","Fine-tune"]) & (samp_all["sampling"]=="None")]
        if sub.empty: continue
        # 分組標題
        ws8.merge_cells(f"A{current_row}:{get_column_letter(len(SCOLS_ZH))}{current_row}")
        set_cell(ws8, current_row, 1, f"▶ 採樣策略：{grp_name}",
                 font=HDR2_FONT, fill=HDR2_FILL, align=LFT_ALIGN)
        ws8.row_dimensions[current_row].height = 22
        current_row += 1
        write_header_row(ws8, current_row, SCOLS_ZH, height=42)
        current_row += 1
        for r_idx, (_, row) in enumerate(sub.iterrows()):
            rf = ds_fills.get(str(row.get("dataset","")), None)
            for c_idx, col in enumerate(SCOLS_EN, 1):
                val = row.get(col, "")
                if isinstance(val, float) and np.isnan(val): val = "-"
                nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
                set_cell(ws8, current_row, c_idx, val, font=NORM_FONT, fill=rf,
                         align=CTR_ALIGN, border=THIN_BORDER, number_format=nf)
            ws8.row_dimensions[current_row].height = 18
            current_row += 1
        current_row += 1
    freeze_and_autofit(ws8)

# ─────────────────────────────────────────────────────────
# Sheet 9: Old vs New 集成比較
# ─────────────────────────────────────────────────────────
# 把各資料集的 ensemble_old_3 / ensemble_new_3 / ensemble_all_6 並排
on_dfs = []
for ds, epath in [
    ("Bankruptcy", os.path.join(BASE,"ensemble","bankruptcy_ensemble_results.csv")),
    ("Stock",      os.path.join(BASE,"ensemble","stock_ensemble_results.csv")),
    ("Medical",    os.path.join(BASE,"ensemble","medical_ensemble_results.csv")),
]:
    df = read_csv_safe(epath)
    if not df.empty:
        df.insert(0, "dataset", ds)
        on_dfs.append(df)
if on_dfs:
    on_all = pd.concat(on_dfs, ignore_index=True)
    on_all = norm_cols(on_all)
    # 只保留三主角
    on_all = on_all[on_all["method"].isin(["ensemble_old_3","ensemble_new_3","ensemble_all_6"])]
    strat, train, samp = [], [], []
    for m in on_all["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    on_all.insert(2, "strategy",   strat)
    on_all.insert(3, "train_data", train)
    on_all.insert(4, "sampling",   samp)
    for col in METRICS_EN:
        if col not in on_all.columns: on_all[col] = np.nan
    OCOLS_ZH = ["資料集", "方法名稱", "策略", "訓練資料"] + METRICS_ZH
    OCOLS_EN = ["dataset", "method", "strategy", "train_data"] + METRICS_EN
    ws9 = wb.create_sheet("09_Old_vs_New_集成")
    ws9.merge_cells(f"A1:{get_column_letter(len(OCOLS_ZH))}1")
    set_cell(ws9, 1, 1, "Old vs New 集成比較 — ensemble_old_3 / ensemble_new_3 / ensemble_all_6 × 3 資料集",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws9.row_dimensions[1].height = 36
    write_header_row(ws9, 2, OCOLS_ZH, height=48)
    ds_fills2 = {"Bankruptcy": PatternFill("solid","FFF2CC"), "Stock": PatternFill("solid","E2EFDA"),
                 "Medical": PatternFill("solid","FCE4D6")}
    on_all = on_all.sort_values(["dataset","method"]).reset_index(drop=True)
    ds_r = 3
    for r_idx, (_, row) in enumerate(on_all.iterrows()):
        er = ds_r + r_idx
        rf = ds_fills2.get(str(row.get("dataset","")), None)
        for c_idx, col in enumerate(OCOLS_EN, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
            set_cell(ws9, er, c_idx, val, font=NORM_FONT, fill=rf, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws9.row_dimensions[er].height = 18
    de = ds_r + len(on_all) - 1
    for ci, col in enumerate(METRICS_EN, 5):
        add_color_scale(ws9, get_column_letter(ci), ds_r, de, col in ("Type1_Error","Type2_Error"))
    freeze_and_autofit(ws9)

# ─────────────────────────────────────────────────────────
# Sheet 10: Multi-Seed 穩定性
# ─────────────────────────────────────────────────────────
ms_dfs = []
for ds, fpath in [
    ("Bankruptcy", os.path.join(BASE,"multi_seed","bankruptcy_multi_seed.csv")),
    ("Stock",      os.path.join(BASE,"multi_seed","stock_multi_seed.csv")),
    ("Medical",    os.path.join(BASE,"multi_seed","medical_multi_seed.csv")),
]:
    df = read_csv_safe(fpath)
    if not df.empty:
        df.insert(0, "dataset", ds)
        ms_dfs.append(df)
if ms_dfs:
    ms_all = pd.concat(ms_dfs, ignore_index=True)
    ms_all = norm_cols(ms_all)
    MSCOLS = ["dataset","method","AUC_mean","AUC_std","F1_mean","F1_std",
              "G_Mean_mean","G_Mean_std","Recall_mean","Recall_std",
              "Precision_mean","Precision_std","Type1_Error_mean","Type1_Error_std",
              "Type2_Error_mean","Type2_Error_std"]
    MSCOLS_ZH = ["資料集","方法名稱","AUC\nmean","AUC\nstd","F1\nmean","F1\nstd",
                 "G-Mean\nmean","G-Mean\nstd","Recall★\nmean","Recall★\nstd",
                 "Precision\nmean","Precision\nstd","Type1_Err★\nmean","Type1_Err★\nstd",
                 "Type2_Err\nmean","Type2_Err\nstd"]
    ws10 = wb.create_sheet("10_Multi_Seed穩定性")
    ws10.merge_cells(f"A1:{get_column_letter(len(MSCOLS))}1")
    set_cell(ws10, 1, 1, "多種子穩定性分析（seeds={42,123,456}）— 驗證結果可重現性",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws10.row_dimensions[1].height = 36
    write_header_row(ws10, 2, MSCOLS_ZH, height=48)
    ds_fills3 = {"Bankruptcy": PatternFill("solid","FFF2CC"), "Stock": PatternFill("solid","E2EFDA"),
                 "Medical": PatternFill("solid","FCE4D6")}
    for r_idx, (_, row) in enumerate(ms_all.iterrows()):
        er = 3 + r_idx
        rf = ds_fills3.get(str(row.get("dataset","")), None)
        for c_idx, col in enumerate(MSCOLS, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if isinstance(val, float) and col not in ("dataset","method") else None
            set_cell(ws10, er, c_idx, val, font=NORM_FONT, fill=rf, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws10.row_dimensions[er].height = 18
    freeze_and_autofit(ws10)

# ─────────────────────────────────────────────────────────
# Sheet 11: Feature Selection
# ─────────────────────────────────────────────────────────
fs_adv = read_csv_safe(os.path.join(BASE,"feature_study","fs_advanced_comparison.csv"))
if not fs_adv.empty:
    fs_adv = norm_cols(fs_adv)
    strat, train, samp = [], [], []
    for m in fs_adv["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    fs_adv.insert(2, "strategy",   strat)
    fs_adv.insert(3, "train_data", train)
    fs_adv.insert(4, "sampling",   samp)
    FSCOLS_ZH = ["資料集","FS方法","特徵數","方法名稱","策略","訓練資料","採樣策略"] + METRICS_ZH
    FSCOLS_EN = ["dataset","fs_method","n_features","method","strategy","train_data","sampling"] + METRICS_EN
    for col in METRICS_EN:
        if col not in fs_adv.columns: fs_adv[col] = np.nan
    ws11 = wb.create_sheet("11_特徵選擇FS")
    ws11.merge_cells(f"A1:{get_column_letter(len(FSCOLS_ZH))}1")
    set_cell(ws11, 1, 1, "進階特徵選擇比較 — none / kbest_f / mutual_info / SHAP × 3 資料集",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws11.row_dimensions[1].height = 36
    write_header_row(ws11, 2, FSCOLS_ZH, height=48)
    ds_fills4 = {"bankruptcy": PatternFill("solid","FFF2CC"), "stock": PatternFill("solid","E2EFDA"),
                 "medical": PatternFill("solid","FCE4D6")}
    for r_idx, (_, row) in enumerate(fs_adv.iterrows()):
        er = 3 + r_idx
        rf = ds_fills4.get(str(row.get("dataset","")).lower(), None)
        for c_idx, col in enumerate(FSCOLS_EN, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
            set_cell(ws11, er, c_idx, val, font=NORM_FONT, fill=rf, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws11.row_dimensions[er].height = 18
    de = 3 + len(fs_adv) - 1
    for ci, col in enumerate(METRICS_EN, 8):
        add_color_scale(ws11, get_column_letter(ci), 3, de, col in ("Type1_Error","Type2_Error"))
    freeze_and_autofit(ws11)

# ─────────────────────────────────────────────────────────
# Sheet 12: Base Learner
# ─────────────────────────────────────────────────────────
bl_df = read_csv_safe(os.path.join(BASE,"base_learner","base_learner_comparison.csv"))
if not bl_df.empty:
    bl_df = norm_cols(bl_df)
    BLCOLS_ZH = ["資料集","學習器","方法名稱","策略","訓練資料","採樣策略"] + METRICS_ZH
    BLCOLS_EN = ["dataset","learner","method","strategy","train_data","sampling"] + METRICS_EN
    strat, train, samp = [], [], []
    for m in bl_df["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    bl_df.insert(3, "strategy",   strat)
    bl_df.insert(4, "train_data", train)
    bl_df.insert(5, "sampling",   samp)
    for col in METRICS_EN:
        if col not in bl_df.columns: bl_df[col] = np.nan
    ws12 = wb.create_sheet("12_基礎學習器比較")
    ws12.merge_cells(f"A1:{get_column_letter(len(BLCOLS_ZH))}1")
    set_cell(ws12, 1, 1, "基礎學習器比較 — LightGBM vs XGBoost vs RandomForest × 3 資料集",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws12.row_dimensions[1].height = 36
    write_header_row(ws12, 2, BLCOLS_ZH, height=48)
    l_fills = {"LightGBM": PatternFill("solid","E2EFDA"), "XGBoost": PatternFill("solid","FFF2CC"),
               "RandomForest": PatternFill("solid","FCE4D6")}
    bl_df = bl_df.sort_values(["dataset","learner","method"]).reset_index(drop=True)
    for r_idx, (_, row) in enumerate(bl_df.iterrows()):
        er = 3 + r_idx
        rf = l_fills.get(str(row.get("learner","")), None)
        for c_idx, col in enumerate(BLCOLS_EN, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
            set_cell(ws12, er, c_idx, val, font=NORM_FONT, fill=rf, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws12.row_dimensions[er].height = 18
    de = 3 + len(bl_df) - 1
    for ci, col in enumerate(METRICS_EN, 7):
        add_color_scale(ws12, get_column_letter(ci), 3, de, col in ("Type1_Error","Type2_Error"))
    freeze_and_autofit(ws12)

# ─────────────────────────────────────────────────────────
# Sheet 13: Split Comparison
# ─────────────────────────────────────────────────────────
sp_df = read_csv_safe(os.path.join(BASE,"baseline","bankruptcy_split_comparison.csv"))
if not sp_df.empty:
    sp_df = norm_cols(sp_df)
    SPCOLS_ZH = ["切割方式","方法名稱","策略","訓練資料","採樣策略"] + METRICS_ZH
    SPCOLS_EN = ["split_mode","method","strategy","train_data","sampling"] + METRICS_EN
    strat, train, samp = [], [], []
    for m in sp_df["method"]:
        s, t, sa = get_meta(m)
        strat.append(s); train.append(t); samp.append(sa)
    sp_df.insert(2, "strategy",   strat)
    sp_df.insert(3, "train_data", train)
    sp_df.insert(4, "sampling",   samp)
    for col in METRICS_EN:
        if col not in sp_df.columns: sp_df[col] = np.nan
    ws13 = wb.create_sheet("13_切割方式比較")
    ws13.merge_cells(f"A1:{get_column_letter(len(SPCOLS_ZH))}1")
    set_cell(ws13, 1, 1, "Bankruptcy 切割方式比較 — Chronological（年份）vs Block-CV（位置）",
             font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=CTR_ALIGN)
    ws13.row_dimensions[1].height = 36
    write_header_row(ws13, 2, SPCOLS_ZH, height=48)
    sp_fills = {"chronological": PatternFill("solid","E2EFDA"), "block_cv": PatternFill("solid","FCE4D6")}
    sp_df = sp_df.sort_values(["split_mode","method"]).reset_index(drop=True)
    for r_idx, (_, row) in enumerate(sp_df.iterrows()):
        er = 3 + r_idx
        rf = sp_fills.get(str(row.get("split_mode","")).lower(), None)
        for c_idx, col in enumerate(SPCOLS_EN, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = "-"
            nf = "0.0000" if col in set(METRICS_EN) and isinstance(val, float) else None
            set_cell(ws13, er, c_idx, val, font=NORM_FONT, fill=rf, align=CTR_ALIGN,
                     border=THIN_BORDER, number_format=nf)
        ws13.row_dimensions[er].height = 18
    de = 3 + len(sp_df) - 1
    for ci, col in enumerate(METRICS_EN, 5):
        add_color_scale(ws13, get_column_letter(ci), 3, de, col in ("Type1_Error","Type2_Error"))
    freeze_and_autofit(ws13)

# ─────────────────────────────────────────────────────────
# 儲存
# ─────────────────────────────────────────────────────────
out = os.path.join(BASE, "advisor_summary_20260302.xlsx")
wb.save(out)
print(f"[Done] Saved: {out}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")

